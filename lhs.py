import xml.etree.ElementTree as ET
from openpyxl import load_workbook, Workbook
import datetime
import os
import numpy as np
from scipy.stats import qmc

def parse_col():
    tree = ET.parse("reports/cross.coltest.xml")
    root = tree.getroot()
    collision_number = 0
    
    for h in root.iter("collision"):
        collision_number += 1
        collision_element = ET.SubElement(
            root, "crash",
            type=str(h.attrib["type"]),
            collider=str(h.attrib["collider"]),
            victim=str(h.attrib["victim"])
        )
        collision_element.tail = "    \n    "

    tree.write("reports/summarized_report.xml")
    return collision_number

def parse_fuel():
    tree = ET.parse("reports/emissions.xml")
    root = tree.getroot()

    total_fuel = {}
    count = {}

    for vehicle in root.iter("vehicle"):
        vehicle_id = vehicle.attrib["id"]
        fuel_value = float(vehicle.attrib["fuel"])

        if vehicle_id not in total_fuel:
            total_fuel[vehicle_id] = 0
            count[vehicle_id] = 0
        total_fuel[vehicle_id] += fuel_value
        count[vehicle_id] += 1

    average_fuel = {}
    for vehicle_id in total_fuel:
        average_fuel[vehicle_id] = total_fuel[vehicle_id] / count[vehicle_id]

    for vehicle_id, avg_fuel in average_fuel.items():
        print(f"ID {vehicle_id}: Average Fuel = {avg_fuel:.2f}")

    file_path = "reports/summarized_report.xml"
    tree = ET.parse(file_path)
    root = tree.getroot()

    for vehicle_id, avg_fuel in average_fuel.items():
        vehicle_element = ET.SubElement(
            root, "vehicle",
            id=str(vehicle_id),
            fuel=str(avg_fuel),
            times=str(count[vehicle_id]),
            vehicle_loaded=str(len(total_fuel))
        )
        vehicle_element.tail = "    \n    "

    tree.write(file_path)
    return vehicle_id, avg_fuel, count, len(total_fuel), average_fuel

def create_report(parameters):
    col_num = parse_col()
    vehicle_id, avg_fuel, count, loaded_vehicle, average_fuel_array = parse_fuel()

    ids = [vehicle_id for vehicle_id, avg_fuel in average_fuel_array.items()]
    fuels = [avg_fuel for vehicle_id, avg_fuel in average_fuel_array.items()]
    times = [count[vehicle_id] for vehicle_id, avg_fuel in average_fuel_array.items()]

    file_name = 'reports/summarized_report.xlsx'
    if os.path.exists(file_name):
        workbook = load_workbook(file_name)
    else:
        workbook = Workbook()
        workbook.remove(workbook.active)

    timestamp = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    worksheet_name = f'Sheet_{timestamp}'
    worksheet = workbook.create_sheet(title=worksheet_name)

    worksheet.append(['vehicle_id', 'avg_fuel', 'times', 'col num', 'parameters'])

    for i, (id, fuel, time) in enumerate(zip(ids, fuels, times), start=2):
        worksheet.cell(row=i, column=1, value=id)
        worksheet.cell(row=i, column=2, value=fuel)
        worksheet.cell(row=i, column=3, value=time)

    worksheet['D2'] = col_num

    parameter_titles = [
        "t", "max_speed", "lower_acc", "upper_acc", "speed-loc_fac",
        "reaction_t", "safety_distance", "vehicle_length", "edge_len",
        "raw_intersection_num", "column_intersection_num", "detector_pos",
        "node_num", "MAX_ITER", "TIME_GAP", "TOLERANCE", "RHO"
    ]
    for i in range(2, 19):
        worksheet.cell(row=i, column=5, value=parameter_titles[i-2])
        worksheet.cell(row=i, column=6, value=parameters[i-2])

    workbook.save(file_name)

# ======================
# LHS Sampling Entegrasyonu
# ======================
def run_lhs_experiments(n_samples=10):
    # Değişken parametrelerin alt-üst sınırları
    param_bounds = [
        (60, 240),    # max_speed
        (-6, -1),     # lower_acc
        (1, 6),       # upper_acc
        (0.0, 0.5),   # speed_loc_fac
        (0.2, 3.0),   # reaction_t
        (0, 12),      # safety_distance
        (2, 6),       # vehicle_length
        (1, 10),      # MAX_ITER
        (0.5, 6.0),   # TIME_GAP
        (0.01, 15.0), # TOLERANCE
        (0.1, 10.0)   # RHO
    ]

    sampler = qmc.LatinHypercube(d=len(param_bounds))
    lhs_samples = sampler.random(n=n_samples)
    lhs_scaled = qmc.scale(lhs_samples,
                           [low for (low, high) in param_bounds],
                           [high for (low, high) in param_bounds])

    for sample in lhs_scaled:
        # Sabit parametreler:
        params = [
            1,                      # t
            sample[0],              # max_speed
            sample[1],              # lower_acc
            sample[2],              # upper_acc
            sample[3],              # speed_loc_fac
            sample[4],              # reaction_t
            sample[5],              # safety_distance
            sample[6],              # vehicle_length
            500,                    # edge_len
            3,                      # raw_intersection_num
            3,                      # column_intersection_num
            2,                      # detector_pos
            50,                     # node_num
            sample[7],              # MAX_ITER
            sample[8],              # TIME_GAP
            sample[9],              # TOLERANCE
            sample[10]              # RHO
        ]
        create_report(params)

if __name__ == "__main__":
    run_lhs_experiments(n_samples=5)  # örnek olarak 5 senaryo
