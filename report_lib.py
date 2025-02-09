import xml.etree.ElementTree as ET

import array as arr

import xlsxwriter
import random

from openpyxl import load_workbook, Workbook
import datetime
import os

def parse_col():
    tree = ET.parse("reports/cross.coltest.xml")
    root = tree.getroot()
    collision_number = 0
    
    for h in root.iter("collision"):
        if h.attrib["type"] == "junction":
            collision_number += 1
            collision_element = ET.SubElement(root, "crash", type=str(h.attrib["type"]), collider=str(h.attrib["collider"]), victim=str(h.attrib["victim"]))
            collision_element.tail= "    \n    "
        else:
            pass

    tree.write("reports/summarized_report.xml")

    return collision_number

def parse_fuel():
    tree = ET.parse("reports/emissions.xml")
    root = tree.getroot()

    # Initialize dictionaries to store total fuel and count for each ID
    total_fuel = {}
    count = {}

    # Iterate through the XML data
    for vehicle in root.iter("vehicle"):
        vehicle_id = vehicle.attrib["id"]
        fuel_value = float(vehicle.attrib["fuel"])

        # Update total fuel and count for the current ID
        if vehicle_id not in total_fuel:
            total_fuel[vehicle_id] = 0
            count[vehicle_id] = 0
        total_fuel[vehicle_id] += fuel_value
        count[vehicle_id] += 1

    # Calculate average fuel for each ID
    average_fuel = {}
    for vehicle_id in total_fuel:
        average_fuel[vehicle_id] = total_fuel[vehicle_id] / count[vehicle_id]

    for vehicle_id, avg_fuel in average_fuel.items():
        print(f"ID {vehicle_id}: Average Fuel = {avg_fuel:.2f}")

    file_path = "reports/summarized_report.xml"
    # Check if the file exists
    
    tree = ET.parse(file_path)
    root = tree.getroot()

    # Add vehicle elements to the XML
    for vehicle_id, avg_fuel in average_fuel.items():
        vehicle_element = ET.SubElement(root, "vehicle", id=str(vehicle_id), fuel=str(avg_fuel), times=str(count[vehicle_id]), vehicle_loaded=str(len(total_fuel)))
        #vehicle_element.text = str(avg_fuel)
        vehicle_element.tail= "    \n    "

    # Write the changes back to the file
    tree.write(file_path)

    return vehicle_id, avg_fuel, count, len(total_fuel), average_fuel

def create_report(parameters):

    col_num = parse_col()
    vehicle_id, avg_fuel, count, loaded_vehicle, average_fuel_array = parse_fuel()
    # Example data
    # Try to do as much processing outside of initializing the workbook
    # Everything beetween Workbook() and close() gets trapped in an exception
    ids = [vehicle_id for vehicle_id, avg_fuel in average_fuel_array.items()]
    fuels = [avg_fuel for vehicle_id, avg_fuel in average_fuel_array.items()]
    times = [count[vehicle_id] for vehicle_id, avg_fuel in average_fuel_array.items()]
    # Data location inside excel
    data_start_loc = [1, 0] # xlsxwriter rquires list, no tuple
    #data_end_loc = [data_start_loc[0] + len(ids), 0]
    data_start_loc2 = [1, 1]
    data_start_loc3 = [1, 2]
    #workbook = xlsxwriter.Workbook('reports/summarized_report.xlsx')
    file_name = 'reports/summarized_report.xlsx'
    # # Charts are independent of worksheets
    # chart = workbook.add_chart({'type': 'line'})
    # chart.set_y_axis({'name': 'Random jiggly bit values'})
    # chart.set_x_axis({'name': 'Sequential order'})
    # chart.set_title({'name': 'Insecure randomly jiggly bits'})

    #worksheet = workbook.add_worksheet()

    # Check if the file already exists
    if os.path.exists(file_name):
        # Load the existing workbook
        workbook = load_workbook(file_name)
    else:
        # Create a new workbook
        workbook = Workbook()
        workbook.remove(workbook.active)  # Remove the default sheet


    # Generate a unique worksheet name
    timestamp = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    worksheet_name = f'Sheet_{timestamp}'

    # Add a new worksheet with the unique name
    worksheet = workbook.create_sheet(title=worksheet_name)

    # Write the first row
    worksheet.append(['vehicle_id', 'avg_fuel', 'times', 'col num', 'parameters'])

    # Write the data columns
    for i, (id, fuel, time) in enumerate(zip(ids, fuels, times), start=2):
        worksheet.cell(row=i, column=1, value=id)
        worksheet.cell(row=i, column=2, value=fuel)
        worksheet.cell(row=i, column=3, value=time)

    worksheet['D2'] = col_num

    parameter_titles = ["t", "max_speed", "lower_acc", "upper_acc", "speed-loc_fac", "reaction_t", "safety_distance", "vehicle_length"]
    for i in range(2,9):
        worksheet.cell(row=i, column=5, value=parameter_titles[i-2])
        worksheet.cell(row=i, column=6, value=parameters[i-2])

    # Save the workbook
    workbook.save(file_name)

    #workbook.close()  # Write to file
    pass